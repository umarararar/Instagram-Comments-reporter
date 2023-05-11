import instaloader
from time import sleep
import undetected_chromedriver as uc
from selenium.webdriver.chrome.options import Options
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from time import sleep
from selenium.common.exceptions import StaleElementReferenceException
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook("book.xlsx")

# Select the sheet that contains the usernames and passwords
sheet = wb.active

###############


def scrollComments():
    while True:
        try:

            # Get the height of the comment section
            height = driver.execute_script("return arguments[0].scrollHeight", comment_section)
            # Scroll to the bottom of the comment section
            #print(height)
            driver.execute_script("arguments[0].scrollTo(0, arguments[0].scrollHeight)", comment_section)
           # sleep(1) # Wait for the page to load
            wait = WebDriverWait(driver, 5)
            # load_more_xpath = "/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/section/main/div[1]/div[1]/article/div/div[2]/div/div[2]/div[1]/ul/li/div/button"
            load_more_xpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/li/div/button"

            #/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/li/div/button
            load_more_button = wait.until(EC.element_to_be_clickable((By.XPATH, load_more_xpath)))
            load_more_button.click()
            
        except:
            driver.execute_script("arguments[0].scrollTo(0, 0)", comment_section)
            break
#reporting from different accounts..different account se login krky report
def otherAccountsReport(driver, instaLinks):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username = row[0]
        password = row[1]
        print(f"Logging in with username: {username} and password: {password}")
        driver.get("https://www.instagram.com/")
        

        try:
            #/html/body/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/button[1]
            cookies = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/button[1]")))
            cookies.click()
        except:
            pass
        username_input = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[name='username']")))
        password_input = WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[name='password']")))

        username_input.send_keys(username)
        password_input.send_keys(password)

        login_button = driver.find_element(By.XPATH,"//button[@type='submit']")
        login_button.click()
        try:
            notnow=WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[3]/button[2]")))
            notnow.click()
        except:
            pass
        try:
            Homebutton = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[1]/div/div/div/div/div[2]/div[1]/div/div/a/div")))
            Homebutton.click()
        except:
             print("\nLogin failed for this account. Maybe the sheet has wrong credentials for this account or maybe it is banned!\n")
             continue
        
        sleep(1)

        # open a new tab
        #driver.execute_script("window.open();")

        # switch to the new tab
        #driver.switch_to.window(driver.window_handles[1])

        # navigate to the URL in the new tab, logging in with all ids and reporting
        for l in instaLinks:
            driver.get(l)
            wait = WebDriverWait(driver, 30)
            #"")
           # threeXpath = "/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div[2]/section/main/div/header/section/div[1]/div[2]/button"
            try:
                three = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div[2]/section/main/div/header/section/div[1]/div[2]/button")))
                three.click()
                    
                newReportXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div/button[3]"
                newReport = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, newReportXpath)))
                newReport.click()
                
                accountXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div[3]/button[2]/div/div[1]"
                account = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, accountXpath)))
                account.click()

                reasonXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div[1]/button[1]/div/div[1]"
                reason = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, reasonXpath)))
                reason.click()
                
                bullyXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div[1]/button[8]/div/div[1]"
                bully = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, bullyXpath)))
                bully.click()

                meXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div[3]/button[1]/div/div[1]"
                mex = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, meXpath)))
                mex.click()

                Submit2Xpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/button"
                submit2 = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, Submit2Xpath)))
                submit2.click()

                close2Xpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div/div/div/div[4]/button"
                close2 = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, close2Xpath)))
                close2.click()
            except:
                continue
            # perform some action on the new tab, e.g. search for a keyword
            #
            # wait for some time to see the search results
            sleep(5)
           # driver.close()
        more=WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[1]/div/div/div/div/div[3]/div/a/div")))
        more.click()
        sleep(1)
        logout=WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[2]/div/div/div[1]/div[1]/div/div/div/div[1]/div/div/div[1]/div/div[6]/div[1]/div/div/div/div/div")))
        logout.click()
        sleep(1)
        
        try:
            buttonl=WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/button[1]")))
            buttonl.click()
        except:
            pass
        try:
            switch=WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#mount_0_0_bG > div > div > div.x9f619.x1n2onr6.x1ja2u2z > div > div > div > div.x78zum5.xdt5ytf.x10cihs4.x1t2pt76.x1n2onr6.x1ja2u2z > section > main > article > div._ab1y > div:nth-child(2) > div > div.xdj266r.x11i5rnm.xat24cr.x1mh8g0r > div:nth-child(1)")))
            switch.click()
        except:
            pass

    # close the new tab
   # driver.close()

    # switch back to the main tab
    #driver.switch_to.window(driver.window_handles[0])


def loginFromFirstAccountForReporting():
    for row in sheet.iter_rows(min_row=3, max_row=3, values_only=True):
        username = row[0]
        password = row[1]
        print(f"Logging in with username: {username} and password: {password}")

        

        username_input = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[name='username']")))
        password_input = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"input[name='password']")))

        username_input.send_keys(username)
        password_input.send_keys(password)

        login_button = driver.find_element(By.XPATH,"//button[@type='submit']")
        login_button.click()
        try:
            Homebutton = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[1]/div/div/div/div/div[2]/div[1]/div/div/a/div")))
            Homebutton.click()
        except:
             print("\nLogin failed for this account. Maybe the sheet has wrong credentials for this account or maybe it is banned!\n")
             continue
        sleep(7)


# Create an instance of Instaloader and login to the account
# L = instaloader.Instaloader()
# L.interactive_login("shakirjani2024")
words=input("Enter the word you want to search for: ")
# Get the profile of the account
#profName=input("Enter name of profile to be searched: ")
# profile = instaloader.Profile.from_username(L.context, profName)

# Loop through the posts of the profile
#for post in profile.get_posts():
    # Get the comments of the post

options = uc.ChromeOptions()
options.add_argument("--incognito")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--disable-cookies")
driver = uc.Chrome(options=options)
actions = ActionChains(driver)
driver.maximize_window()


driver.get('https://www.instagram.com/')
            
try:
    cookiesb = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/button[1]")))
    cookiesb.click()
except:
    pass
sleep(1)


##########

#login from first account for reporting the post
loginFromFirstAccountForReporting()

with open('followers_output.txt', 'r') as f:
    lines = f.readlines()

for line in lines:
    profName = line.strip()  # remove newline character
    driver.get('https://www.instagram.com/' + profName)
    sleep(3)
    try:
        checkPrivate=WebDriverWait(driver, 7).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div[2]/section/main/div/div/article/div[1]/div/h2")))
        driver.execute_script("arguments[0].scrollIntoView();", checkPrivate)
        print(checkPrivate.text,"!\n")
        continue
    except:
        try:
            dontExist=WebDriverWait(driver, 7).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/section/main/div/div/span")))
            driver.execute_script("arguments[0].scrollIntoView();", dontExist)
            print(dontExist.text,"!\n")
            continue
        except:
            pass



    postCount=WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div[2]/section/main/div/header/section/ul/li[1]/span/span")))
    post_count=int(postCount.text)
    sleep(3)

    c=0

    real_post_count=int(post_count/3)
    flag=False
    counterPost=0
    for i in range(1,real_post_count+1):
        c+=1
        for j in range(1,4):
            
            if(i*j>=30):
                flag=True
                print("!!!!!!!")
                break
            else:
                try:
                    post=WebDriverWait(driver, 7).until(EC.element_to_be_clickable((By.XPATH,f"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div[2]/section/main/div/div[3]/article/div[1]/div/div[{i}]/div[{j}]/a/div/div[2]")))
                    driver.execute_script("arguments[0].scrollIntoView();", post)
                    post.click()
                except:
                    try:
                        post1=WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH,f"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div[2]/section/main/div/div[2]/article/div/div/div[{i}]/div[{j}]")))
                        driver.execute_script("arguments[0].scrollIntoView();", post1)
                        post1.click()
                    except:
                        continue
                       # /html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div[2]/section/main/div/div[2]/article/div/div/div[1]/div[1]
                postUrl=driver.current_url  # print the current URL
                counterPost+=1

                #########################################
                ##########################################
                # print(postUrl.rsplit('/', 2)[-2])
                # pp=postUrl.rsplit('/', 2)[-2]
                # post = instaloader.Post.from_shortcode(L.context, pp)
    

            
            # post = instaloader.Post.from_url(L.context, postUrl)


    ####################with code comments
            #if there is no comment on post
                try:
                    wait55 = WebDriverWait(driver, 10)
                    comment_section = wait55.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul')))
                    scrollComments()
                    comxpath = f"/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/ul/div/li/div/div/div[2]"
                
                    comments = wait55.until(EC.presence_of_all_elements_located((By.XPATH, comxpath)))
                except:
                    cross=WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div")))
                    cross.click()
                    continue
                
                wait = WebDriverWait(driver, 50)
                reportUsers = []
                print("\nrow#",c,"\n")
                for comment in comments:
                    com= comment.find_element(By.TAG_NAME,"span").text
                    if words in com:
                        username = comment.find_element(By.TAG_NAME,"h3").text
                    # com= comment.find_element(By.TAG_NAME,"span").text
                        reportUsers.append(username)
                        print("username: ",username," Comment: ",com)

                    #sleep(1)
                print("Commentors to be reported: ",reportUsers)



    #             ######with instaloader comments

    #             comments = post.get_comments()
    #             reportUsers = []
            
    #         #  Loop through the comments and print the required information if the comment contains the desired keyword
    #             for comment in comments:
    #                 if words in comment.text:
    #                     sleep(1)
    #                     username = comment.owner.username
    #                     reportUsers.append(username)
    #                     print("#", comment.id, "Name:", username, "\t Comment:", comment.text)

    #             post_url = "https://www.instagram.com/p/" + post.shortcode + "/"   
            
            
            
    #              ######with instaloader comments
                #if no comment with specific keyword, go to next post
                if not reportUsers:
                    cross=WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div")))
                    cross.click()
                    print("##############Can't find comment with given keyword on post: ",postUrl)
                    continue
                
                else:
                    print("COMMENT FOUND ON POST: ",postUrl)

    


    #             waitc = WebDriverWait(driver, 30)
    #          #   comment_section = waitc.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/section/main/div[1]/div[1]/article/div/div[2]/div/div[2]/div[1]/ul')))
    # #/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul
    #             comment_section = waitc.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul')))

    #         #scroll and load all the commments
    
                

    #         #    xpath = f"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/section/main/div[1]/div[1]/article/div/div[2]/div/div[2]/div[1]/ul/ul/div/li/div/div/div[2]/h3/div/div/div/a"
    # #/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/ul/div/li/div/div/div[2]/h3
                xpath = f"/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/ul/div/li/div/div/div[2]/h3"

    #             wait = WebDriverWait(driver, 30)

                names = WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located((By.XPATH, xpath)))

                wait = WebDriverWait(driver, 50)
                instaLinks=[]

            #Links of ids to be reported, stored in instaLinks list
                for name in names:
                    try:
                        if not name.text:
                            break
                        if name and name.text in reportUsers:
                            link = "https://www.instagram.com/" + name.text
                            instaLinks.append(link)
                        
                    except StaleElementReferenceException:
                        pass
                print("Links of commentors' ID to be reported: ",instaLinks)

    #             # for l in instaLinks:    
    #             #     print(l)

                count1=0

                for name in names:
                    count1 += 1
                    try:
                        if not name.text:
                            break
                        if name and name.text in reportUsers:
                        # link = "https://www.instagram.com/" + name.text
                            #badCommentXpath = "/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/section/main/div[1]/div[1]/article/div/div[2]/div/div[2]/div[1]/ul/ul[{}]/div/li/div/div/div[2]/div[1]/span".format(count1)
                            badCommentXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/ul[{}]/div/li/div/div/div[2]/div[1]/span".format(count1)

                        #   /html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/ul[7]/div/li/div/div/div[2]/div[1]/span
                            
                            badComment = wait.until(EC.element_to_be_clickable((By.XPATH, badCommentXpath)))
                            print(reportUsers,"###",name.text,"########")
                            actions.move_to_element(badComment).perform()
                            sleep(1)
                        #  dotsXpath = "/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/section/main/div[1]/div[1]/article/div/div[2]/div/div[2]/div[1]/ul/ul[{}]/div/li/div/div/div[2]/div[2]/span/div[2]/div/button".format(count1)
                            dotsXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/ul[{}]/div/li/div/div/div[2]/div[2]/span/div[2]/div/button/div/div".format(count1)
    
                        #/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[3]/div/div/div/div/div[2]/div/article/div/div[2]/div/div/div[2]/div[1]/ul/ul[3]/div/li/div/div/div[2]/div[2]/span/div[2]/div/button
                            dots = wait.until(EC.element_to_be_clickable((By.XPATH, dotsXpath)))
                            dots.click()
                            sleep(1)
                            #/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div/button[1]
                            #reportXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div/button[1]"
                            reportXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div/button[1]"
                            reportButton = wait.until(EC.element_to_be_clickable((By.XPATH, reportXpath)))
                            reportButton.click()
                            sleep(1)
                            try:
                                #/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div[3]/button[3]/div
                            # hateXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div[3]/button[3]/div/div[1]"
                                harassmentXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div[3]/button[6]/div"
                                harassment = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, harassmentXpath)))
                                harassment.click()
                                sleep(1)
                                whoXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/div/div[3]/button[1]/div"
                                whobutton = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, whoXpath)))
                                whobutton.click()
                                sleep(1)
                                #/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/button
                                #submitXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/button"
                                submitXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/div/div/button"
                                submitButton = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, submitXpath)))
                                submitButton.click()
                                sleep(1)

                                #/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div/div/div/div[4]/button
                                #closeXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div/div/div/div[4]/button"
                                closeXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div/div/div/div[4]/button"
                                closeButton = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, closeXpath)))
                                closeButton.click()

                            except:
                                try:
                                    #/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[1]/div/div[3]/div/button 
                                    crossXpath = "/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[1]/div/div[3]/div/button"
                                    crossButton = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, crossXpath)))
                                    crossButton.click()
                                except:
                                    pass
                    except StaleElementReferenceException:
                        pass
                
                cross=WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div")))
                cross.click()
                more=WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[1]/div[1]/div/div/div/div/div[3]/div/a/div")))
                more.click()
                sleep(1)
                logout=WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div/div[2]/div/div/div[1]/div[1]/div/div/div/div[1]/div/div/div[1]/div/div[6]/div[1]/div/div/div/div/div")))
                logout.click()
                sleep(1)
                
                try:
                    buttonl=WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[3]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/div/div[2]/button[1]")))
                    buttonl.click()
                except:
                    pass
                try:
                    switch=WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.CSS_SELECTOR,"#mount_0_0_bG > div > div > div.x9f619.x1n2onr6.x1ja2u2z > div > div > div > div.x78zum5.xdt5ytf.x10cihs4.x1t2pt76.x1n2onr6.x1ja2u2z > section > main > article > div._ab1y > div:nth-child(2) > div > div.xdj266r.x11i5rnm.xat24cr.x1mh8g0r > div:nth-child(1)")))
                    switch.click()
                except:
                    pass
                #reporting the user with different accounts
                #driver.close()

                otherAccountsReport(driver, instaLinks)
                print('##########################################################################################')
                driver.get('https://www.instagram.com/')

                
                try:
                    cookiesb = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div[2]/div/div/div[2]/div/div/div[1]/div/div[2]/div/div/div/div/div[2]/div/button[1]")))
                    cookiesb.click()
                except:
                    pass
                sleep(1)

                loginFromFirstAccountForReporting()
                driver.get('https://www.instagram.com/'+profName)

                ########################################
                ############################################

            #  sleep(3)
                
        if(flag==True):
            print(counterPost)
            break



print("Thank you for using our bot")



sleep(200)
